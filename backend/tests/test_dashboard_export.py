import io
import os
import tempfile

import openpyxl
import pytest
from openpyxl import Workbook

from app.excel_to_parquet import excel_to_parquet
from app.query_engine import run_export_query
from app.routers.dashboard import DETAIL_COLUMN_LABELS, GROUP_AGG_LABELS, GROUP_BY_LABELS, rows_to_xlsx_bytes

HEADERS = [
    "Mã đơn hàng", "Ngày đặt hàng", "Trạng Thái Đơn Hàng", "Lý do hủy",
    "SKU phân loại hàng", "Tên sản phẩm", "Tên phân loại hàng",
    "Giá gốc", "Số lượng", "Số lượng sản phẩm được hoàn trả",
]
ROWS = [
    ["O1", "2026-02-01 00:01", "Đã hủy", "Giao hàng thất bại", "A100-1", "SP A", "Áo", 100000, 2, 0],
    ["O2", "2026-02-02 09:00", "Đã hủy", "Người mua đổi ý", "B200-2", "SP B", "Quần", 50000, 3, 0],
    ["O3", "2026-02-03 10:00", "Hoàn thành", "", "C300-1", "SP C", "Áo", 20000, 4, 4],
]


@pytest.fixture
def parquet_path():
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    parquet_bytes, row_count, _ = excel_to_parquet(buf)
    assert row_count == 3

    fd, path = tempfile.mkstemp(suffix=".parquet")
    with os.fdopen(fd, "wb") as f:
        f.write(parquet_bytes)
    yield path
    os.remove(path)


def test_export_ungrouped_xlsx_has_expected_headers_and_rows(parquet_path):
    rows = run_export_query(parquet_path)
    col_keys = ["orderId", "product", "doanhSo"]
    buf = rows_to_xlsx_bytes(rows, col_keys, DETAIL_COLUMN_LABELS)

    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == ["Mã đơn hàng", "Sản phẩm", "Doanh số"]
    assert ws.max_row == 1 + len(rows)  # header + every matching row, unpaginated
    assert {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)} == {"O1", "O2", "O3"}


def test_export_grouped_xlsx_has_dynamic_group_label_header(parquet_path):
    rows = run_export_query(parquet_path, group_by="category")
    col_keys = ["groupValue", "rowCount", "doanhSo"]
    labels = {**GROUP_AGG_LABELS, "groupValue": GROUP_BY_LABELS["category"]}
    buf = rows_to_xlsx_bytes(rows, col_keys, labels)

    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == ["Danh mục", "Số dòng", "Doanh số"]
    assert ws.max_row == 1 + 2  # Áo, Quần


def test_export_empty_rows_still_produces_valid_workbook_with_headers():
    buf = rows_to_xlsx_bytes([], ["orderId", "doanhSo"], DETAIL_COLUMN_LABELS)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == ["Mã đơn hàng", "Doanh số"]
    assert ws.max_row == 1
