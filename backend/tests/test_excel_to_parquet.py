import io
import re
import zipfile

import pyarrow.parquet as pq
from openpyxl import Workbook

from app.excel_to_parquet import MappingError, excel_to_parquet, get_original_headers, read_excel_rows

HEADERS = [
    "Mã đơn hàng", "Ngày đặt hàng", "Trạng Thái Đơn Hàng", "Lý do hủy",
    "SKU phân loại hàng", "Tên sản phẩm", "Giá gốc", "Số lượng",
    "Số lượng sản phẩm được hoàn trả",
]

ROWS = [
    ["O1", "2026-02-01 00:01", "Đã hủy", "Giao hàng thất bại", "A100-1", "SP A", 100000, 2, 0],
    ["O2", "2026-02-02 09:00", "Đã hủy", "Người mua đổi ý", "B200-2", "SP B", 50000, 3, 0],
    ["O3", "2026-02-03 10:00", "Hoàn thành", "", "C300-1", "SP C", 20000, 4, 4],
    ["O4", "2026-02-04 11:00", "Hoàn thành", "", "D400-3", "SP D", 30000, 5, 2],
    ["O5", "2026-02-05 12:00", "Hoàn thành", "", "E500-1", "SP E", 200000, 1, 0],
    ["O6", "2026-02-06 13:00", "Đang giao hàng", "", "F600-2", "SP F", 80000, 2, 0],
]


def make_xlsx_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_excel_to_parquet_end_to_end():
    parquet_bytes, row_count, mapping = excel_to_parquet(make_xlsx_bytes())

    assert row_count == 6
    assert mapping["date"] == "Ngày đặt hàng"
    assert mapping["status"] == "Trạng Thái Đơn Hàng"
    assert mapping["originalPrice"] == "Giá gốc"
    assert mapping["orderId"] == "Mã đơn hàng"

    table = pq.read_table(io.BytesIO(parquet_bytes))
    df = table.to_pandas()
    assert len(df) == 6

    total_doanh_so = df["doanhSo"].sum()
    assert total_doanh_so == 940000

    by_status = df.groupby("trangThai")["doanhSo"].sum().to_dict()
    assert by_status["Hủy sau XK"] == 200000
    assert by_status["Hủy chưa XK"] == 150000
    assert by_status["Hoàn hàng"] == 80000
    assert by_status["Hoàn 1 phần"] == 150000
    gmv = by_status.get("Hoàn thành", 0) + by_status.get("Đang giao", 0)
    assert gmv == 360000

    # skuVariant survives the full pipeline; the parent "sku" code is
    # deliberately not persisted (query_engine.py recomputes it from
    # skuVariant at query time — see test_derive.py and test_query_engine.py).
    row_a = df[df["orderId"] == "O1"].iloc[0]
    assert row_a["skuVariant"] == "A100-1"
    assert "sku" not in df.columns


def test_missing_date_column_raises():
    wb = Workbook()
    ws = wb.active
    ws.append(["Sản phẩm", "Số lượng", "Giá gốc"])
    ws.append(["A", 1, 1000])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    try:
        excel_to_parquet(buf)
        assert False, "expected MappingError"
    except MappingError:
        pass


def test_get_original_headers():
    headers = get_original_headers(make_xlsx_bytes())
    assert headers == HEADERS


def _make_one_row_tag_per_cell_xlsx_bytes():
    """Builds an .xlsx whose worksheet XML has ONE <row> wrapper per
    individual cell instead of one <row> per actual spreadsheet row (seen
    verbatim in a real TikTok Shop order export). openpyxl's read_only
    mode streams by <row> element and ends up keeping just a single
    column's worth of cells per real row instead of raising anything —
    this reproduces that malformed structure from a normal openpyxl file.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Order ID", "Order Status", "Quantity"])
    ws.append(["O1", "Đã hủy", 2])
    ws.append(["O2", "Hoàn thành", 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    malformed_sheet_data = (
        "<sheetData>"
        '<row r="1"><c r="A1" t="str"><v>Order ID</v></c></row>'
        '<row r="1"><c r="B1" t="str"><v>Order Status</v></c></row>'
        '<row r="1"><c r="C1" t="str"><v>Quantity</v></c></row>'
        '<row r="2"><c r="A2" t="str"><v>O1</v></c></row>'
        '<row r="2"><c r="B2" t="str"><v>Đã hủy</v></c></row>'
        '<row r="2"><c r="C2" t="n"><v>2</v></c></row>'
        '<row r="3"><c r="A3" t="str"><v>O2</v></c></row>'
        '<row r="3"><c r="B3" t="str"><v>Hoàn thành</v></c></row>'
        '<row r="3"><c r="C3" t="n"><v>1</v></c></row>'
        "</sheetData>"
    )
    out = io.BytesIO()
    with zipfile.ZipFile(buf) as zin, zipfile.ZipFile(out, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = re.sub(r"<sheetData>.*</sheetData>", malformed_sheet_data, data.decode("utf-8"), flags=re.S)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    out.seek(0)
    return out


def test_read_excel_rows_recovers_from_one_row_tag_per_cell_export():
    # Regression: a malformed export (one <row> XML wrapper per cell
    # instead of per real row) made openpyxl's fast read_only mode
    # silently collapse every row down to just its first column, with no
    # exception raised — read_excel_rows must detect the suspiciously
    # narrow (<=1 column) result and retry without that fast path.
    rows, headers = read_excel_rows(_make_one_row_tag_per_cell_xlsx_bytes())
    assert headers == ["Order ID", "Order Status", "Quantity"]
    assert len(rows) == 2
    assert rows[0]["Order ID"] == "O1"
    assert rows[0]["Order Status"] == "Đã hủy"
    assert rows[0]["Quantity"] == 2
    assert rows[1]["Order ID"] == "O2"


def test_mapping_override_reconverts_with_chosen_columns():
    # Regression for the "Chỉnh cột" PATCH endpoint: since the Parquet's
    # columns are fixed at conversion time, overriding the mapping must
    # actually reconvert the file with the admin's chosen columns — not
    # just relabel stored metadata. Here we deliberately map "product" to
    # the SKU column instead of "Tên sản phẩm" to prove the override wins
    # over auto-detection.
    override = {
        "date": "Ngày đặt hàng",
        "orderId": "Mã đơn hàng",
        "product": "SKU phân loại hàng",  # would never be auto-detected as product
        "quantity": "Số lượng",
        "price": "Giá gốc",
        "originalPrice": "Giá gốc",
        "returnedQty": "Số lượng sản phẩm được hoàn trả",
        "status": "Trạng Thái Đơn Hàng",
        "cancelReason": "Lý do hủy",
    }
    parquet_bytes, row_count, mapping = excel_to_parquet(make_xlsx_bytes(), mapping_override=override)
    assert row_count == 6
    assert mapping["product"] == "SKU phân loại hàng"

    table = pq.read_table(io.BytesIO(parquet_bytes))
    df = table.to_pandas()
    row_a = df[df["orderId"] == "O1"].iloc[0]
    assert row_a["product"] == "A100-1"  # took the overridden column, not "SP A"


def test_mapping_override_missing_date_still_raises():
    try:
        excel_to_parquet(make_xlsx_bytes(), mapping_override={"product": "Tên sản phẩm"})
        assert False, "expected MappingError"
    except MappingError:
        pass


DISCOUNT_HEADERS = HEADERS + ["Người bán trợ giá", "Mã giảm giá của Shop", "Số tiền người mua thanh toán"]

DISCOUNT_ROWS = [
    # MULTI1: 2-line order. Shopee repeats the order-level shop voucher
    # (10.000) on every line; it must be prorated by each line's share of
    # "Số tiền người mua thanh toán" (300.000 + 700.000 = 1.000.000 total).
    ["MULTI1", "2026-02-01 00:01", "Hoàn thành", "", "A100-1", "SP A", 150000, 2, 0, 4000, 10000, 300000],
    ["MULTI1", "2026-02-01 00:01", "Hoàn thành", "", "B200-1", "SP B", 700000, 1, 0, 6000, 10000, 700000],
    # SINGLE1: 1-line order -> ratio must be exactly 100%.
    ["SINGLE1", "2026-02-02 09:00", "Hoàn thành", "", "C300-1", "SP C", 50000, 1, 0, 1000, 5000, 50000],
]


def make_discount_xlsx_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"
    ws.append(DISCOUNT_HEADERS)
    for r in DISCOUNT_ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_discount_and_voucher_prorated_across_multi_line_order():
    parquet_bytes, row_count, mapping = excel_to_parquet(make_discount_xlsx_bytes())
    assert row_count == 3
    assert mapping["sellerSubsidy"] == "Người bán trợ giá"
    assert mapping["shopVoucher"] == "Mã giảm giá của Shop"
    assert mapping["buyerPaidAmount"] == "Số tiền người mua thanh toán"

    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()

    line1 = df[(df["orderId"] == "MULTI1") & (df["skuVariant"] == "A100-1")].iloc[0]
    line2 = df[(df["orderId"] == "MULTI1") & (df["skuVariant"] == "B200-1")].iloc[0]
    single = df[df["orderId"] == "SINGLE1"].iloc[0]

    # discount = (Người bán trợ giá / Số lượng) x Số lượng thực — no returns
    # here so Số lượng thực == Số lượng.
    assert line1["discount"] == 4000 / 2 * 2
    assert line2["discount"] == 6000 / 1 * 1

    # line1 is 300.000/1.000.000 = 30% of the order's paid amount; line2 is 70%.
    assert line1["voucher"] == 10000 * 0.3 / 2 * 2
    assert line2["voucher"] == 10000 * 0.7 / 1 * 1

    # Single-line order -> ratio is exactly 100%, so voucher == the shop's
    # full voucher value for that line (no proration needed).
    assert single["voucher"] == 5000
    assert single["discount"] == 1000

    # orderPaidRatio is persisted per row (not just used transiently) so the
    # query-time Phí AFF join can reuse the exact same proration.
    assert line1["orderPaidRatio"] == 0.3
    assert line2["orderPaidRatio"] == 0.7
    assert single["orderPaidRatio"] == 1.0


def test_discount_and_voucher_default_to_zero_when_columns_absent():
    # Existing Reports converted before this feature has no "Người bán trợ
    # giá" / "Mã giảm giá của Shop" / "Số tiền người mua thanh toán"
    # columns — must not error, discount/voucher should just be 0.
    parquet_bytes, row_count, mapping = excel_to_parquet(make_xlsx_bytes())
    assert "sellerSubsidy" not in mapping
    assert "shopVoucher" not in mapping
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    assert (df["discount"] == 0).all()
    assert (df["voucher"] == 0).all()


FEE_HEADERS = DISCOUNT_HEADERS + ["Phí cố định", "Phí dịch vụ", "Phí xử lý giao dịch"]

FEE_ROWS = [
    # F1: 2-line order. Fee columns hold the SAME order-level total on both
    # lines (1000+2000+500=3500), prorated by paid-amount ratio (40/60).
    ["F1", "2026-02-01 00:01", "Hoàn thành", "", "A100-1", "SP A", 100000, 2, 0, 4000, 10000, 400000, 1000, 2000, 500],
    ["F1", "2026-02-01 00:01", "Hoàn thành", "", "B200-1", "SP B", 150000, 1, 0, 1500, 10000, 600000, 1000, 2000, 500],
    # F2: single-line order -> ratio 100%, and the only (first) line for Piship.
    ["F2", "2026-02-02 09:00", "Hoàn thành", "", "C300-1", "SP C", 50000, 1, 0, 1000, 5000, 50000, 500, 300, 200],
]


def make_fee_xlsx_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"
    ws.append(FEE_HEADERS)
    for r in FEE_ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_platform_fee_prorated_and_piship_assigned_to_first_line_only():
    parquet_bytes, row_count, mapping = excel_to_parquet(make_fee_xlsx_bytes())
    assert row_count == 3
    assert mapping["fixedFee"] == "Phí cố định"
    assert mapping["serviceFee"] == "Phí dịch vụ"
    assert mapping["transactionFee"] == "Phí xử lý giao dịch"

    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    line1 = df[(df["orderId"] == "F1") & (df["skuVariant"] == "A100-1")].iloc[0]
    line2 = df[(df["orderId"] == "F1") & (df["skuVariant"] == "B200-1")].iloc[0]
    single = df[df["orderId"] == "F2"].iloc[0]

    # (1000+2000+500)=3500 total order fee, prorated 40%/60% by paid amount.
    assert line1["platformFee"] == 3500 * 0.4
    assert line2["platformFee"] == 3500 * 0.6
    assert single["platformFee"] == 1000  # 500+300+200, ratio 100%

    # Piship (1.620/order, flat) goes to only the first surviving line.
    assert line1["piship"] == 1620
    assert line2["piship"] == 0
    assert single["piship"] == 1620


def test_piship_gated_by_sales_channel():
    # Piship is Shopee's own delivery-partner fee — must not apply when a
    # non-Shopee channel is selected at upload time (e.g. TikTok, which
    # has no Piship-equivalent at all, per the user).
    for channel, expect_piship in [(None, True), ("SHOPEE", True), ("Shopee", True), ("TikTok Shop", False), ("Lazada", False)]:
        parquet_bytes, _, _ = excel_to_parquet(make_fee_xlsx_bytes(), sales_channel_name=channel)
        df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
        first_line_piship = df[(df["orderId"] == "F1") & (df["skuVariant"] == "A100-1")].iloc[0]["piship"]
        assert first_line_piship == (1620 if expect_piship else 0), f"channel={channel!r}"


def test_piship_rate_changes_on_23_05_2026():
    # Confirmed with the user 2026-09-03: Shopee raised Piship from 1.620
    # to 2.700 starting exactly 23/05/2026, compared against the order's
    # own "Ngày đặt hàng" — not upload date, not today's date.
    rows = [
        ["P1", "2026-05-22 10:00", "Hoàn thành", "", "A100-1", "SP A", 100000, 1, 0],  # day before
        ["P2", "2026-05-23 10:00", "Hoàn thành", "", "B200-1", "SP B", 100000, 1, 0],  # change date itself
        ["P3", "2026-09-03 10:00", "Hoàn thành", "", "C300-1", "SP C", 100000, 1, 0],  # well after
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parquet_bytes, _, _ = excel_to_parquet(buf)
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    by_order = df.set_index("orderId")
    assert by_order.loc["P1", "piship"] == 1620
    assert by_order.loc["P2", "piship"] == 2700
    assert by_order.loc["P3", "piship"] == 2700


def test_missing_orderid_column_raises():
    # "Mã đơn hàng" is required: Piship (flat fee per order, first line
    # only) and Voucher/Phí sàn proration both depend on being able to
    # group rows into orders — without it there's no way to compute these
    # correctly, so the upload is rejected rather than guessed at.
    override = {
        "date": "Ngày đặt hàng",
        "price": "Giá gốc",
        "quantity": "Số lượng",
        "originalPrice": "Giá gốc",
        "returnedQty": "Số lượng sản phẩm được hoàn trả",
        "status": "Trạng Thái Đơn Hàng",
        "cancelReason": "Lý do hủy",
        "skuVariant": "SKU phân loại hàng",
        "product": "Tên sản phẩm",
        # "orderId" deliberately omitted.
    }
    try:
        excel_to_parquet(make_xlsx_bytes(), mapping_override=override)
        assert False, "expected MappingError"
    except MappingError as e:
        assert "Mã đơn hàng" in str(e)


def test_piship_assigned_only_to_first_line_of_each_order():
    parquet_bytes, row_count, mapping = excel_to_parquet(make_xlsx_bytes())
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    # ROWS above is 6 distinct single-line orders (O1..O6) -> each is its
    # own order's first (and only) line, so each still gets 1.620.
    assert (df["piship"] == 1620).all()


def test_missing_required_column_lists_all_missing_fields():
    # Every field marked required=True in mapping.FIELDS must be present;
    # a file missing several of them should name all of them, not just
    # the first one found.
    try:
        excel_to_parquet(make_xlsx_bytes(), mapping_override={"date": "Ngày đặt hàng"})
        assert False, "expected MappingError"
    except MappingError as e:
        msg = str(e)
        for label in ["Mã đơn hàng", "Số lượng", "SL sản phẩm hoàn trả"]:
            assert label in msg, f"expected {label!r} in error message: {msg!r}"
        # No "status" mapped at all -> falls to the revenue-or-originalPrice
        # check (see test_..._when_status_absent below), not the Shopee/
        # TikTok-shaped Trạng thái/Lý do hủy requirement.
        assert "Giá gốc (hoặc Doanh thu" in msg


def test_missing_original_price_and_cancel_reason_required_when_status_mapped():
    # A Shopee/TikTok-shaped file (tracks order status) must still carry
    # originalPrice/cancelReason — status-based derivation depends on them,
    # so silently defaulting them would mis-classify cancelled/returned
    # orders. Confirmed with the user 2026-08-28 alongside the new
    # no-status in-house/POS file shape that relaxed these two.
    try:
        excel_to_parquet(make_xlsx_bytes(), mapping_override={
            "date": "Ngày đặt hàng", "orderId": "Mã đơn hàng", "quantity": "Số lượng",
            "returnedQty": "Số lượng sản phẩm được hoàn trả", "status": "Trạng Thái Đơn Hàng",
        })
        assert False, "expected MappingError"
    except MappingError as e:
        msg = str(e)
        assert "Giá gốc" in msg
        assert "Lý do hủy" in msg


def test_original_price_not_required_when_status_absent_but_revenue_mapped():
    # The in-house POS/social/web/Zalo file shape (no Trạng thái đơn hàng,
    # no Giá gốc — confirmed with the user 2026-08-28) must still convert
    # successfully when "Doanh thu" is mapped instead.
    parquet_bytes, row_count, mapping = excel_to_parquet(make_xlsx_bytes(), mapping_override={
        "date": "Ngày đặt hàng", "orderId": "Mã đơn hàng", "quantity": "Số lượng",
        "returnedQty": "Số lượng sản phẩm được hoàn trả", "revenue": "Giá gốc",
    })
    assert row_count > 0
    assert "status" not in mapping
    assert "originalPrice" not in mapping

    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    by_order = df.set_index("orderId")
    # No status column at all -> "huỷ"/"đang giao" text-detection never
    # fires (O1/O2's real "Đã hủy" status text is simply never read), but
    # the return-quantity-based branches are still independent of status
    # and still apply: O3 (returnedQty==quantity) -> "Hoàn hàng", O4
    # (partial return) -> "Hoàn 1 phần", everything else -> "Hoàn thành"
    # (no cancellation concept for this file shape — see
    # derive_order_status's status_known param).
    assert by_order.loc["O1", "trangThai"] == "Hoàn thành"
    assert by_order.loc["O3", "trangThai"] == "Hoàn hàng"
    assert by_order.loc["O4", "trangThai"] == "Hoàn 1 phần"
    # No originalPrice mapped, but revenue is -> doanhSo falls back to the
    # raw "Doanh thu" (here "Giá gốc") value directly, not 0.
    assert by_order.loc["O1", "doanhSo"] == 100000  # ROWS[0]'s "Giá gốc" value


def test_voucher_and_platform_fee_fall_back_to_quantity_share_when_buyerpaidamount_not_mapped():
    # Regression: when "Số tiền người mua thanh toán" isn't mapped,
    # order_paid_ratio used to always be 0.0 (dividing by a total that's
    # never populated), silently zeroing Voucher and Phí sàn for the whole
    # report even though the source columns had real values. It should
    # instead fall back to prorating by each line's share of the order's
    # total quantity.
    override = {
        "date": "Ngày đặt hàng",
        "orderId": "Mã đơn hàng",
        "price": "Giá gốc",
        "quantity": "Số lượng",
        "originalPrice": "Giá gốc",
        "returnedQty": "Số lượng sản phẩm được hoàn trả",
        "status": "Trạng Thái Đơn Hàng",
        "cancelReason": "Lý do hủy",
        "skuVariant": "SKU phân loại hàng",
        "product": "Tên sản phẩm",
        "sellerSubsidy": "Người bán trợ giá",
        "shopVoucher": "Mã giảm giá của Shop",
        "fixedFee": "Phí cố định",
        "serviceFee": "Phí dịch vụ",
        "transactionFee": "Phí xử lý giao dịch",
        # "buyerPaidAmount" deliberately omitted.
    }
    parquet_bytes, row_count, mapping = excel_to_parquet(make_fee_xlsx_bytes(), mapping_override=override)
    assert "buyerPaidAmount" not in mapping
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()

    line1 = df[(df["orderId"] == "F1") & (df["skuVariant"] == "A100-1")].iloc[0]  # quantity 2
    line2 = df[(df["orderId"] == "F1") & (df["skuVariant"] == "B200-1")].iloc[0]  # quantity 1
    single = df[df["orderId"] == "F2"].iloc[0]  # quantity 1, only line in its order

    # F1's 2 lines split 2:1 by quantity (order total quantity = 3) instead
    # of collapsing to 0.
    assert line1["orderPaidRatio"] == 2 / 3
    assert line2["orderPaidRatio"] == 1 / 3
    assert line1["platformFee"] == 3500 * (2 / 3)
    assert line2["platformFee"] == 3500 * (1 / 3)
    assert line1["voucher"] == 10000 * (2 / 3)
    assert line2["voucher"] == 10000 * (1 / 3)

    # A single-line order still gets ratio 1.0 either way.
    assert single["orderPaidRatio"] == 1.0
    assert single["platformFee"] == 1000
    assert single["voucher"] == 5000


def test_platform_fee_and_piship_default_to_zero_when_columns_absent():
    parquet_bytes, row_count, mapping = excel_to_parquet(make_xlsx_bytes())
    assert "fixedFee" not in mapping
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    assert (df["platformFee"] == 0).all()
    # Piship is still assigned per-order regardless of whether fee columns
    # exist — one row per distinct orderId in this dataset (O1..O6) should
    # each get 1620 (all single-line orders here).
    assert (df["piship"] == 1620).all()


# TikTok-only, optional columns feeding the Dashboard's "Kênh nhỏ"
# classification (see query_engine._aff_channel_join) — skuId is TikTok's
# internal numeric SKU id, a real ~19-digit value that would lose precision
# if it round-tripped through a float instead of staying plain text.
TIKTOK_KENH_NHO_HEADERS = HEADERS + ["SKU ID", "Creator Handle", "Order Channel"]
TIKTOK_KENH_NHO_ROWS = [
    ROWS[0] + ["1730315401307982614", "bbstores.vn", "Videos"],
    ROWS[1] + ["", "", ""],
]


def make_tiktok_kenh_nho_xlsx_bytes():
    wb = Workbook()
    ws = wb.active
    ws.append(TIKTOK_KENH_NHO_HEADERS)
    for r in TIKTOK_KENH_NHO_ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_sku_id_creator_handle_content_channel_are_captured_as_plain_text():
    parquet_bytes, _, mapping = excel_to_parquet(make_tiktok_kenh_nho_xlsx_bytes())
    assert mapping["skuId"] == "SKU ID"
    assert mapping["creatorHandle"] == "Creator Handle"
    assert mapping["contentChannel"] == "Order Channel"

    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    row1 = df[df["orderId"] == "O1"].iloc[0]
    # Exact string match, not "1.7303154013079826e+18" or similar — a float
    # round-trip would silently corrupt this real TikTok SKU id.
    assert row1["skuId"] == "1730315401307982614"
    assert row1["creatorHandle"] == "bbstores.vn"
    assert row1["contentChannel"] == "Videos"

    row2 = df[df["orderId"] == "O2"].iloc[0]
    assert row2["skuId"] == ""
    assert row2["creatorHandle"] == ""
    assert row2["contentChannel"] == ""


def test_sku_id_absent_when_column_not_in_file():
    parquet_bytes, _, mapping = excel_to_parquet(make_xlsx_bytes())
    assert "skuId" not in mapping
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    assert (df["skuId"] == "").all()


# ---- Combined 31 LVS/HARA/WEBSITE/ZALO in-house file (real headers/values
# confirmed against sale_report_28_08_2026_927871_1.xlsx, 2026-08-28) — one
# file mixes all 4 channels, marked per-row by "Kênh bán hàng", and has no
# Trạng thái/Giá gốc/Lý do hủy at all. Also the real source of the
# negative-sign discovery: "Số sản phẩm trả"/"Giảm giá"/"Hoàn trả" are all
# stored NEGATIVE in this file, unlike Shopee/TikTok's positive convention.
COMBINED_CHANNEL_HEADERS = [
    "Ngày", "Kênh bán hàng", "SKU", "Mã đơn hàng", "Tên khách hàng",
    "Số sản phẩm", "Số sản phẩm trả", "Doanh thu", "Giảm giá", "Hoàn trả",
]
COMBINED_CHANNEL_ROWS = [
    # POS -> 31 LVS, no return, no discount.
    ["2026-08-01", "POS", "A100-1", "P1", "Chị A", 6, 0, 3894000, 0, 0],
    # Harasocial -> HARA, discount stored negative.
    ["2026-08-02", "Harasocial", "B200-1", "P2", "Chị B", 10, 0, 6490000, -1947000, 0],
    # Web -> WEBSITE.
    ["2026-08-03", "Web", "C300-1", "P3", "Chị C", 2, 0, 1058000, -529000, 0],
    # Zalo (lowercase "zalo" anywhere in the value) -> ZALO, full return:
    # both "Số sản phẩm trả" and "Hoàn trả" stored negative.
    ["2026-08-04", "Zalo", "D400-1", "P4", "Chị D", 1, -1, 1490000, 0, -1490000],
]


def make_combined_channel_xlsx_bytes():
    wb = Workbook()
    ws = wb.active
    ws.append(COMBINED_CHANNEL_HEADERS)
    for r in COMBINED_CHANNEL_ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_combined_channel_file_maps_and_converts():
    parquet_bytes, row_count, mapping = excel_to_parquet(make_combined_channel_xlsx_bytes())
    assert row_count == 4
    assert mapping["channelRaw"] == "Kênh bán hàng"
    assert mapping["discountAmount"] == "Giảm giá"
    assert mapping["refundAmount"] == "Hoàn trả"
    assert "status" not in mapping
    assert "originalPrice" not in mapping


def test_combined_channel_row_maps_to_correct_sales_channel_and_gates_piship():
    parquet_bytes, _, _ = excel_to_parquet(make_combined_channel_xlsx_bytes())
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    by_order = df.set_index("orderId")

    assert by_order.loc["P1", "channelOverride"] == "31 LVS"
    assert by_order.loc["P2", "channelOverride"] == "HARA"
    assert by_order.loc["P3", "channelOverride"] == "WEBSITE"
    assert by_order.loc["P4", "channelOverride"] == "ZALO"
    # None of the 4 combined channels have Piship — must be gated off even
    # though no upload-time channel was picked (sales_channel_name=None
    # would otherwise default Piship on, see channel_has_piship).
    assert (df["piship"] == 0).all()


def test_combined_channel_negative_sign_convention_normalized_to_positive():
    # "Số sản phẩm trả"/"Giảm giá"/"Hoàn trả" are stored NEGATIVE in this
    # file (confirmed against the real export) — must be normalized to
    # positive, matching Shopee/TikTok's convention, not passed through
    # as-is (a raw negative "SL hoàn trả" would corrupt so_luong_thuc's
    # subtraction, and a raw negative discount would backwards-increase
    # doanh thu thuần instead of reducing it).
    parquet_bytes, _, _ = excel_to_parquet(make_combined_channel_xlsx_bytes())
    df = pq.read_table(io.BytesIO(parquet_bytes)).to_pandas()
    by_order = df.set_index("orderId")

    p2 = by_order.loc["P2"]
    assert p2["discount"] == 1947000  # positive, not -1947000
    assert p2["doanhSo"] == 6490000  # = raw "Doanh thu", matching Doanh số = Doanh thu

    p4 = by_order.loc["P4"]
    assert p4["returnedQty"] == 1  # abs(-1), not -1
    assert p4["soLuongThuc"] == 0  # quantity(1) - returnedQty(1) -> fully returned
    assert p4["trangThai"] == "Hoàn hàng"  # return-quantity branch, independent of status_known
    assert p4["hoanAmount"] == 1490000  # abs(-1490000), not -1490000

