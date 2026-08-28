import io
import os
import tempfile
from datetime import datetime

import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook

from app.excel_to_parquet import excel_to_parquet
from app.query_engine import run_export_query, run_grouped_rows_query, run_rows_query, run_summary_query

HEADERS = [
    "Mã đơn hàng", "Ngày đặt hàng", "Trạng Thái Đơn Hàng", "Lý do hủy",
    "SKU phân loại hàng", "Tên sản phẩm", "Tên phân loại hàng",
    "Giá gốc", "Số lượng", "Số lượng sản phẩm được hoàn trả",
]

# Same 6-row dataset as test_excel_to_parquet, plus a category column so
# category filtering/breakdown can be exercised.
ROWS = [
    ["O1", "2026-02-01 00:01", "Đã hủy", "Giao hàng thất bại", "A100-1", "SP A", "Áo", 100000, 2, 0],
    ["O2", "2026-02-02 09:00", "Đã hủy", "Người mua đổi ý", "B200-2", "SP B", "Quần", 50000, 3, 0],
    ["O3", "2026-02-03 10:00", "Hoàn thành", "", "C300-1", "SP C", "Áo", 20000, 4, 4],
    ["O4", "2026-02-04 11:00", "Hoàn thành", "", "D400-3", "SP D", "Quần", 30000, 5, 2],
    ["O5", "2026-02-05 12:00", "Hoàn thành", "", "E500-1", "SP E", "Áo", 200000, 1, 0],
    ["O6", "2026-02-06 13:00", "Đang giao hàng", "", "F600-2", "SP F", "Quần", 80000, 2, 0],
]


@pytest.fixture
def parquet_path():
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parquet_bytes, row_count, mapping = excel_to_parquet(buf)
    assert row_count == 6

    fd, path = tempfile.mkstemp(suffix=".parquet")
    with os.fdopen(fd, "wb") as f:
        f.write(parquet_bytes)
    yield path
    os.remove(path)


def test_summary_kpis_match_expected(parquet_path):
    result = run_summary_query(parquet_path)
    kpis = result["kpis"]

    assert kpis["doanhSo"] == 940000
    # GMV now also includes "Hoàn 1 phần" rows, counted at their net
    # (post-return) amount: O4 = 30000 * soLuongThuc(3) = 90000, added to
    # O5 (200000, Hoàn thành) + O6 (160000, Đang giao) -> 450000.
    assert kpis["gmv"] == 450000
    assert kpis["huyChuaXK"] == 150000
    assert kpis["huySauXK"] == 200000
    # Doanh số hoàn = Giá gốc x SL hoàn trả (not the full line's doanhSo) —
    # O3: 20000*4=80000, O4: 30000*2=60000 -> 140000. This is intentionally
    # NOT part of the "4 KPIs sum to Doanh số" reconciliation any more —
    # the user confirmed dropping that invariant for this KPI.
    assert kpis["hoan"] == 140000
    assert kpis["rowCount"] == 6


def test_summary_hoan_counts_returned_qty_regardless_of_final_status():
    # "Hủy sau XK" wins priority over the return branches in derive_order_status
    # when the order was ALSO cancelled — but Doanh số hoàn (Giá gốc x SL hoàn
    # trả) must still count that row's returned units, since it's not scoped
    # to trangThai at all.
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append(["O1", "2026-02-01 00:01", "Đã hủy", "Giao hàng thất bại", "A100-1", "SP A", "Áo", 100000, 2, 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    parquet_bytes, row_count, _ = excel_to_parquet(buf)
    assert row_count == 1

    fd, path = tempfile.mkstemp(suffix=".parquet")
    with os.fdopen(fd, "wb") as f:
        f.write(parquet_bytes)
    try:
        result = run_summary_query(path)
        assert result["kpis"]["hoan"] == 100000  # 100000 * 1
    finally:
        os.remove(path)


def test_summary_facets_unaffected_by_status_filter(parquet_path):
    result = run_summary_query(parquet_path, status="Hoàn thành")
    # Facets come from the whole report, not the filtered set.
    assert set(result["facets"]["statuses"]) == {
        "Hủy sau XK", "Hủy chưa XK", "Hoàn hàng", "Hoàn 1 phần", "Hoàn thành", "Đang giao",
    }
    assert set(result["facets"]["categories"]) == {"Áo", "Quần"}
    # But the KPI numbers themselves ARE filtered — only O5 has trangThai
    # exactly "Hoàn thành" (O3 is "Hoàn hàng", O4 is "Hoàn 1 phần" — distinct buckets).
    assert result["kpis"]["doanhSo"] == 200000


def test_summary_category_filter(parquet_path):
    result = run_summary_query(parquet_path, category="Áo")
    # O1 (100k*2=200k, Hủy sau XK), O3 (20k*4=80k, Hoàn hàng), O5 (200k*1=200k, Hoàn thành)
    assert result["kpis"]["doanhSo"] == 200000 + 80000 + 200000


def test_summary_to_date_is_inclusive_of_the_whole_day(parquet_path):
    # Regression: "Đến ngày" comes from <input type="date"> as a plain
    # "YYYY-MM-DD" (implicitly midnight). O1 is timestamped "2026-02-01
    # 00:01" — one minute past that midnight — so a naive "date" <= to_date
    # comparison would wrongly exclude it. to_date must mean end-of-day.
    result = run_summary_query(parquet_path, to_date="2026-02-01")
    assert result["kpis"]["rowCount"] == 1
    assert result["kpis"]["doanhSo"] == 200000  # O1 only: 100000 * 2

    # And it must not leak into the next day.
    result_before = run_summary_query(parquet_path, to_date="2026-01-31")
    assert result_before["kpis"]["rowCount"] == 0


def test_top_products_sorted_desc(parquet_path):
    result = run_summary_query(parquet_path)
    values = [p["value"] for p in result["topProducts"]]
    assert values == sorted(values, reverse=True)
    # SP A (100k*2) and SP E (200k*1) are tied at 200.000 for the top spot —
    # which one DuckDB returns first among ties isn't guaranteed, so just
    # confirm the top value itself is correct.
    assert result["topProducts"][0]["value"] == 200000
    assert result["topProducts"][0]["label"] in ("SP A", "SP E")


def test_rows_pagination(parquet_path):
    page1 = run_rows_query(parquet_path, page=1, page_size=4)
    assert page1["total"] == 6
    assert len(page1["rows"]) == 4

    page2 = run_rows_query(parquet_path, page=2, page_size=4)
    assert len(page2["rows"]) == 2

    ids_seen = {r["orderId"] for r in page1["rows"]} | {r["orderId"] for r in page2["rows"]}
    assert ids_seen == {"O1", "O2", "O3", "O4", "O5", "O6"}


def test_rows_search(parquet_path):
    result = run_rows_query(parquet_path, search="sp e")
    assert result["total"] == 1
    assert result["rows"][0]["orderId"] == "O5"


def test_rows_sort_by_doanh_so_desc(parquet_path):
    result = run_rows_query(parquet_path, sort="doanhSo", sort_dir="desc", page_size=10)
    values = [r["doanhSo"] for r in result["rows"]]
    assert values == sorted(values, reverse=True)


def test_rows_sql_injection_attempt_on_sort_is_ignored(parquet_path):
    # Not a real vulnerability test (sort isn't parameterizable in DuckDB),
    # just confirms the whitelist silently falls back instead of erroring
    # or executing arbitrary SQL.
    result = run_rows_query(parquet_path, sort='"date"; DROP TABLE x; --', page_size=10)
    assert result["total"] == 6


# ---- Multi-Report aggregation (Dashboard no longer pins to one Report — see
# the "xem bằng bộ lọc" change: it queries every ready Report's Parquet
# together and the date/category/status filters narrow the combined set) ----

MARCH_HEADERS = HEADERS
MARCH_ROWS = [
    ["O7", "2026-03-01 08:00", "Hoàn thành", "", "G700-1", "SP G", "Áo", 90000, 2, 0],
    ["O8", "2026-03-02 09:00", "Đang giao hàng", "", "H800-1", "SP H", "Quần", 60000, 1, 0],
]


@pytest.fixture
def parquet_path_march():
    wb = Workbook()
    ws = wb.active
    ws.append(MARCH_HEADERS)
    for r in MARCH_ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parquet_bytes, row_count, mapping = excel_to_parquet(buf)
    assert row_count == 2

    fd, path = tempfile.mkstemp(suffix=".parquet")
    with os.fdopen(fd, "wb") as f:
        f.write(parquet_bytes)
    yield path
    os.remove(path)


def test_summary_aggregates_across_multiple_reports(parquet_path, parquet_path_march):
    result = run_summary_query([parquet_path, parquet_path_march])
    # Feb total (940.000) + March total (90000*2 + 60000*1 = 240.000)
    assert result["kpis"]["doanhSo"] == 940000 + 240000
    assert result["kpis"]["rowCount"] == 6 + 2


def test_summary_date_filter_narrows_across_reports(parquet_path, parquet_path_march):
    # Only March rows should count when filtered to March.
    result = run_summary_query([parquet_path, parquet_path_march], from_date="2026-03-01", to_date="2026-03-31")
    assert result["kpis"]["doanhSo"] == 240000
    assert result["kpis"]["rowCount"] == 2


def test_rows_pagination_across_multiple_reports(parquet_path, parquet_path_march):
    result = run_rows_query([parquet_path, parquet_path_march], page_size=100)
    assert result["total"] == 8
    ids_seen = {r["orderId"] for r in result["rows"]}
    assert ids_seen == {"O1", "O2", "O3", "O4", "O5", "O6", "O7", "O8"}


def test_summary_empty_source_list_returns_zeroed_result():
    result = run_summary_query([])
    assert result["kpis"]["doanhSo"] == 0
    assert result["kpis"]["rowCount"] == 0
    assert result["facets"]["categories"] == []


def test_rows_empty_source_list_returns_empty_page():
    result = run_rows_query([], page_size=15)
    assert result["total"] == 0
    assert result["rows"] == []


def test_summary_doanh_thu_thuan_is_gmv_when_no_discount_or_voucher(parquet_path):
    # parquet_path has no "Người bán trợ giá"/"Mã giảm giá của Shop"/"Số
    # tiền người mua thanh toán" columns mapped, so discount/voucher are 0
    # for every row -> Doanh thu thuần should equal GMV exactly.
    result = run_summary_query(parquet_path)
    assert result["kpis"]["doanhThuThuan"] == result["kpis"]["gmv"] == 450000


def test_summary_nmv_also_nets_out_piship_even_without_fee_columns(parquet_path):
    # Piship (1.620/order) doesn't depend on any mapped fee column, but does
    # NOT apply to "Hủy chưa XK" orders (see
    # test_summary_piship_excludes_huy_chua_xk) — parquet_path has 6 orders,
    # one of which (O2) is Hủy chưa XK, so NMV must net out 1.620 x 5.
    result = run_summary_query(parquet_path)
    assert result["kpis"]["nmv"] == result["kpis"]["doanhThuThuan"] - 1620 * 5


def test_summary_piship_excludes_huy_chua_xk(parquet_path):
    # O2 is "Hủy chưa XK" (cancelled before export) — Piship must not apply
    # to it, unlike O1 ("Hủy sau XK", cancelled after export) which still
    # incurs it since shipping already happened.
    result = run_rows_query(parquet_path, page_size=10)
    by_order = {r["orderId"]: r for r in result["rows"]}
    assert by_order["O2"]["piship"] == 0
    assert by_order["O1"]["piship"] == 1620


DISCOUNT_HEADERS = HEADERS + ["Người bán trợ giá", "Mã giảm giá của Shop", "Số tiền người mua thanh toán"]

DISCOUNT_ROWS = [
    # Both rows share Mã đơn hàng "D1" — a single 2-line order. Both are
    # "Hoàn thành" (counts toward GMV/NMV). Shop voucher (10.000) is
    # repeated on both lines and must be prorated 40/60 by paid amount.
    ["D1", "2026-02-01 00:01", "Hoàn thành", "", "A100-1", "SP A", "Áo", 100000, 2, 0, 4000, 10000, 400000],
    ["D1", "2026-02-01 00:01", "Hoàn thành", "", "B200-1", "SP B", "Quần", 150000, 1, 0, 1500, 10000, 600000],
]


@pytest.fixture
def parquet_path_with_discounts():
    wb = Workbook()
    ws = wb.active
    ws.append(DISCOUNT_HEADERS)
    for r in DISCOUNT_ROWS:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parquet_bytes, row_count, mapping = excel_to_parquet(buf)
    assert row_count == 2
    assert "sellerSubsidy" in mapping and "shopVoucher" in mapping

    fd, path = tempfile.mkstemp(suffix=".parquet")
    with os.fdopen(fd, "wb") as f:
        f.write(parquet_bytes)
    yield path
    os.remove(path)


def test_summary_doanh_thu_thuan_nets_out_discount_and_voucher(parquet_path_with_discounts):
    result = run_summary_query(parquet_path_with_discounts)
    kpis = result["kpis"]
    # doanhSo: D1 = 100000*2 = 200000, D2 = 150000*1 = 150000 -> 350000 total, all GMV.
    assert kpis["doanhSo"] == kpis["gmv"] == 350000
    # discount: D1 = 4000/2*2 = 4000, D2 = 1500/1*1 = 1500 -> 5500.
    # voucher: D1 ratio 400000/1000000=0.4 -> 10000*0.4/2*2=4000;
    #          D2 ratio 600000/1000000=0.6 -> 10000*0.6/1*1=6000 -> 10000.
    assert kpis["doanhThuThuan"] == 350000 - 5500 - 10000
    assert kpis["discount"] == 5500
    assert kpis["voucher"] == 10000


def test_summary_nmv_further_nets_out_platform_fee_piship_and_phi_aff(parquet_path_with_discounts):
    # Doanh thu thuần = 334.500 (see test above). D1's 2 lines are one order
    # -> Piship (1.620) applies once, to the first line only. No fee columns
    # mapped in this fixture -> platformFee=0; no cashflow_source -> phiAff=0.
    result = run_summary_query(parquet_path_with_discounts)
    kpis = result["kpis"]
    assert kpis["doanhThuThuan"] == 334500
    assert kpis["platformFee"] == 0
    assert kpis["piship"] == 1620
    assert kpis["phiAff"] == 0
    assert kpis["nmv"] == 334500 - 0 - 1620 - 0


def _write_raw_parquet(rows: list[dict]) -> str:
    """Builds a Parquet file directly (bypassing excel_to_parquet) so the
    schema has exactly the given columns — used to simulate a Report that
    was converted before "discount"/"voucher" existed.
    """
    table = pa.Table.from_pylist(rows)
    fd, path = tempfile.mkstemp(suffix=".parquet")
    os.close(fd)
    pq.write_table(table, path)
    return path


def test_summary_gmv_uses_so_luong_thuc_not_quantity():
    # The real pipeline can never produce a "Hoàn thành"/"Đang giao" row
    # with quantity != soLuongThuc (derive_order_status requires
    # returnedQty == 0 for those statuses) — so this writes a raw Parquet
    # row directly to prove the SQL itself uses "soLuongThuc", not
    # "quantity"/"doanhSo", for GMV.
    rows = [
        {"date": datetime(2026, 2, 1), "orderId": "G1", "sku": "X1",
         "skuVariant": "X1-1", "product": "SP X", "category": "Áo", "customer": "(Không rõ)",
         "quantity": 5.0, "returnedQty": 2.0, "soLuongThuc": 3.0, "price": 0.0, "originalPrice": 10000.0,
         "revenue": 0.0, "doanhSo": 50000.0, "status": "Hoàn thành", "trangThai": "Hoàn thành",
         "discount": 0.0, "voucher": 0.0},
    ]
    path = _write_raw_parquet(rows)
    try:
        result = run_summary_query(path)
        assert result["kpis"]["gmv"] == 30000  # 10000 * 3 (soLuongThuc), not 10000 * 5 (quantity)
    finally:
        os.remove(path)


@pytest.fixture
def old_schema_parquet_path():
    # No "discount"/"voucher" keys at all — mirrors a pre-feature Report.
    rows = [
        {"date": datetime(2026, 2, 1), "orderId": "OLD1", "sku": "X1",
         "skuVariant": "X1-1", "product": "SP X", "category": "Áo", "customer": "(Không rõ)",
         "quantity": 2.0, "returnedQty": 0.0, "soLuongThuc": 2.0, "price": 0.0, "originalPrice": 100000.0,
         "revenue": 0.0, "doanhSo": 200000.0, "status": "Hoàn thành", "trangThai": "Hoàn thành"},
    ]
    path = _write_raw_parquet(rows)
    yield path
    os.remove(path)


def test_summary_old_schema_report_without_discount_columns_does_not_error(old_schema_parquet_path):
    result = run_summary_query(old_schema_parquet_path)
    assert result["kpis"]["doanhSo"] == 200000
    # No discount/voucher columns at all -> NMV falls back to GMV.
    assert result["kpis"]["nmv"] == result["kpis"]["gmv"] == 200000


def test_rows_old_schema_report_without_discount_columns_returns_zero(old_schema_parquet_path):
    result = run_rows_query(old_schema_parquet_path, page_size=10)
    assert result["total"] == 1
    assert result["rows"][0]["discount"] == 0
    assert result["rows"][0]["voucher"] == 0
    assert result["rows"][0]["platformFee"] == 0
    assert result["rows"][0]["piship"] == 0


def test_summary_hoan_falls_back_to_original_price_times_returned_qty_for_old_schema_report():
    # A Report converted before "hoanAmount" existed (no such column at
    # all) — the "hoan" KPI must fall back to recomputing the old formula
    # from originalPrice/returnedQty instead of erroring on a missing
    # column, same backward-compat pattern as discount/voucher.
    rows = [
        {"date": datetime(2026, 2, 1), "orderId": "OLD1", "sku": "X1",
         "skuVariant": "X1-1", "product": "SP X", "category": "Áo", "customer": "(Không rõ)",
         "quantity": 4.0, "returnedQty": 4.0, "soLuongThuc": 0.0, "price": 0.0, "originalPrice": 20000.0,
         "revenue": 0.0, "doanhSo": 80000.0, "status": "Hoàn hàng", "trangThai": "Hoàn hàng"},
    ]
    path = _write_raw_parquet(rows)
    try:
        result = run_summary_query(path)
        assert result["kpis"]["hoan"] == 80000  # 20000 * 4, recomputed
    finally:
        os.remove(path)


def test_summary_mixed_old_and_new_schema_reports_via_union_by_name(
    old_schema_parquet_path, parquet_path_with_discounts
):
    # The Dashboard aggregate endpoint always passes a list of every ready
    # Report's Parquet — some old (no discount/voucher columns), some new.
    # This must not error, and the old Report's rows must contribute 0 to
    # discount/voucher instead of poisoning the whole aggregate with NULLs.
    result = run_summary_query([old_schema_parquet_path, parquet_path_with_discounts])
    assert result["kpis"]["rowCount"] == 1 + 2
    assert result["kpis"]["doanhSo"] == 200000 + 350000
    doanh_thu_thuan = (200000 - 0 - 0) + (350000 - 5500 - 10000)
    assert result["kpis"]["doanhThuThuan"] == doanh_thu_thuan
    # D1's 2-line order -> Piship (1.620) once; OLD1 has no piship column at
    # all (COALESCEs to 0 via union_by_name), not a second order's worth.
    assert result["kpis"]["piship"] == 1620
    assert result["kpis"]["nmv"] == doanh_thu_thuan - 1620


def test_rows_mixed_old_and_new_schema_reports_via_union_by_name(
    old_schema_parquet_path, parquet_path_with_discounts
):
    result = run_rows_query([old_schema_parquet_path, parquet_path_with_discounts], page_size=10)
    assert result["total"] == 3
    by_sku = {r["skuVariant"]: r for r in result["rows"]}
    assert by_sku["X1-1"]["discount"] == 0
    assert by_sku["X1-1"]["voucher"] == 0
    assert by_sku["A100-1"]["discount"] == 4000


# ---- Phí AFF (Cashflow join) ----
# parquet_path_with_discounts has Mã đơn hàng "D1" as a 2-line order:
# line A100-1 (buyerPaid 400.000, ratio 0.4), line B200-1 (buyerPaid
# 600.000, ratio 0.6) — persisted as orderPaidRatio on each row.

def _write_cashflow_parquet(rows: list[dict]) -> str:
    return _write_raw_parquet(rows)


@pytest.fixture
def cashflow_parquet_path():
    path = _write_cashflow_parquet([{"orderId": "D1", "phiAff": 2000.0}])
    yield path
    os.remove(path)


def test_summary_phi_aff_prorated_across_multi_line_order(parquet_path_with_discounts, cashflow_parquet_path):
    result = run_summary_query(parquet_path_with_discounts, cashflow_source=[cashflow_parquet_path])
    # 2000 split 0.4/0.6 across the order's two lines -> sums back to 2000.
    assert result["kpis"]["phiAff"] == 2000


def test_rows_phi_aff_prorated_across_multi_line_order(parquet_path_with_discounts, cashflow_parquet_path):
    result = run_rows_query(parquet_path_with_discounts, page_size=10, cashflow_source=[cashflow_parquet_path])
    by_sku = {r["skuVariant"]: r for r in result["rows"]}
    assert by_sku["A100-1"]["phiAff"] == 2000 * 0.4
    assert by_sku["B200-1"]["phiAff"] == 2000 * 0.6


def test_summary_phi_aff_is_zero_when_no_cashflow_data(parquet_path_with_discounts):
    result = run_summary_query(parquet_path_with_discounts, cashflow_source=None)
    assert result["kpis"]["phiAff"] == 0
    result_empty_list = run_summary_query(parquet_path_with_discounts, cashflow_source=[])
    assert result_empty_list["kpis"]["phiAff"] == 0


def test_rows_phi_aff_is_zero_when_no_cashflow_data(parquet_path_with_discounts):
    result = run_rows_query(parquet_path_with_discounts, page_size=10, cashflow_source=None)
    assert all(r["phiAff"] == 0 for r in result["rows"])


def test_summary_phi_aff_zero_for_orders_report_missing_order_paid_ratio(
    old_schema_parquet_path, cashflow_parquet_path
):
    # old_schema_parquet_path has no "orderPaidRatio" column at all (it
    # predates this feature) — cashflow data existing must not error, and
    # must not misattribute Phí AFF to a report that can't prorate it.
    result = run_summary_query(old_schema_parquet_path, cashflow_source=[cashflow_parquet_path])
    assert result["kpis"]["phiAff"] == 0


def test_cashflow_platform_fee_prorated_and_added_to_orders_file_platform_fee(
    parquet_path_with_discounts,
):
    # TikTok's Cashflow Report also carries a "platformFee" column (Phí
    # sàn), unlike Shopee's (which only ever has "phiAff") — must be
    # prorated by orderPaidRatio the same way Phí AFF already is, and
    # summed with whatever the Orders file itself contributed (0 here,
    # since parquet_path_with_discounts has no fee columns mapped).
    cashflow_path = _write_cashflow_parquet([{"orderId": "D1", "phiAff": 2000.0, "platformFee": 3000.0}])
    try:
        summary = run_summary_query(parquet_path_with_discounts, cashflow_source=[cashflow_path])
        assert summary["kpis"]["platformFee"] == 3000
        assert summary["kpis"]["phiAff"] == 2000

        rows = run_rows_query(parquet_path_with_discounts, page_size=10, cashflow_source=[cashflow_path])
        by_sku = {r["skuVariant"]: r for r in rows["rows"]}
        assert by_sku["A100-1"]["platformFee"] == 3000 * 0.4
        assert by_sku["B200-1"]["platformFee"] == 3000 * 0.6
    finally:
        os.remove(cashflow_path)


def test_cashflow_multiple_rows_for_same_order_are_summed(parquet_path_with_discounts):
    # A single order can appear on more than one row within the SAME
    # Cashflow Report file (TikTok's "income" export does this for a
    # partial return: an original-charge row plus a partial-reversal row
    # sharing the same Mã đơn hàng) — both must be summed via the join's
    # GROUP BY "orderId", not just one row taken. Values mirror a real
    # TikTok order (verified line-by-line against the source formulas
    # against a real income export, 2026-08-27).
    cashflow_path = _write_cashflow_parquet([
        {"orderId": "D1", "phiAff": 0.0, "platformFee": -25602.0},
        {"orderId": "D1", "phiAff": 26294.0, "platformFee": 52302.0},
    ])
    try:
        summary = run_summary_query(parquet_path_with_discounts, cashflow_source=[cashflow_path])
        assert summary["kpis"]["phiAff"] == 26294
        assert summary["kpis"]["platformFee"] == 26700
    finally:
        os.remove(cashflow_path)


def test_cashflow_platform_fee_absent_when_no_cashflow_report_has_it(parquet_path_with_discounts):
    # A Shopee-shaped Cashflow Report (only "phiAff", no "platformFee" at
    # all) must not error — platformFee stays whatever the Orders file
    # itself contributed (0 here).
    cashflow_path = _write_cashflow_parquet([{"orderId": "D1", "phiAff": 2000.0}])
    try:
        result = run_summary_query(parquet_path_with_discounts, cashflow_source=[cashflow_path])
        assert result["kpis"]["platformFee"] == 0
    finally:
        os.remove(cashflow_path)


def test_summary_phi_aff_summed_across_duplicate_orderid_in_multiple_cashflow_reports(
    parquet_path_with_discounts,
):
    # The same Mã đơn hàng "D1" appears in two different ready Cashflow
    # Reports — must be summed once (via GROUP BY) before the join, not
    # double-counted per Orders line.
    path_a = _write_cashflow_parquet([{"orderId": "D1", "phiAff": 2000.0}])
    path_b = _write_cashflow_parquet([{"orderId": "D1", "phiAff": 500.0}])
    try:
        result = run_summary_query(
            parquet_path_with_discounts, cashflow_source=[path_a, path_b],
        )
        assert result["kpis"]["phiAff"] == 2500
    finally:
        os.remove(path_a)
        os.remove(path_b)


# ---- Combo explosion ----
# parquet_path_with_discounts's "D1" order has 2 lines: A100-1 (first line of
# the order, so it carries the order's 1.620 Piship; doanhSo=200000,
# discount=4000, voucher=4000) and B200-1 (second line, Piship=0;
# doanhSo=150000, discount=1500, voucher=6000). A100-1 is set up as a combo
# matching two sub-SKUs (0.5/0.5); B200-1 has no combo match.

def _write_combo_parquet(rows: list[dict]) -> str:
    return _write_raw_parquet(rows)


@pytest.fixture
def combo_parquet_path():
    path = _write_combo_parquet([
        {"skuCombo": "A100-1", "subSku": "X1-1", "ratio": 0.5, "slot": 1},
        {"skuCombo": "A100-1", "subSku": "X2-1", "ratio": 0.5, "slot": 2},
    ])
    yield path
    os.remove(path)


def test_rows_combo_explodes_matching_sku_into_scaled_children(parquet_path_with_discounts, combo_parquet_path):
    result = run_rows_query(parquet_path_with_discounts, page_size=10, combo_source=[combo_parquet_path])
    assert result["total"] == 3  # A100-1 -> 2 children + B200-1 unchanged
    by_sku = {r["skuVariant"]: r for r in result["rows"]}

    assert set(by_sku) == {"X1-1", "X2-1", "B200-1"}

    # Scaled fields: doanhSo/discount/voucher halved for each 0.5 child.
    assert by_sku["X1-1"]["doanhSo"] == 100000
    assert by_sku["X2-1"]["doanhSo"] == 100000
    assert by_sku["X1-1"]["discount"] == 2000
    assert by_sku["X2-1"]["discount"] == 2000
    assert by_sku["X1-1"]["voucher"] == 2000
    assert by_sku["X2-1"]["voucher"] == 2000

    # Unscaled fields: quantity/returnedQty/soLuongThuc copied as-is.
    assert by_sku["X1-1"]["quantity"] == by_sku["X2-1"]["quantity"] == 2
    assert by_sku["X1-1"]["soLuongThuc"] == by_sku["X2-1"]["soLuongThuc"] == 2

    # Piship: A100-1 was D1's first line (1.620) -> all of it goes to the
    # slot-1 child (X1-1); the slot-2 child (X2-1) gets 0, not scaled by ratio.
    assert by_sku["X1-1"]["piship"] == 1620
    assert by_sku["X2-1"]["piship"] == 0

    # B200-1 has no combo match -> passes through completely unchanged.
    assert by_sku["B200-1"]["doanhSo"] == 150000
    assert by_sku["B200-1"]["discount"] == 1500
    assert by_sku["B200-1"]["voucher"] == 6000
    assert by_sku["B200-1"]["piship"] == 0

    # "sku" (parent code) is re-derived from the new skuVariant.
    assert by_sku["X1-1"]["sku"] == "X1"
    assert by_sku["X2-1"]["sku"] == "X2"


def test_summary_combo_explosion_preserves_totals(parquet_path_with_discounts, combo_parquet_path):
    # Splitting a row into scaled children must not change the aggregate
    # totals — they should reconcile exactly with the un-exploded result.
    without_combo = run_summary_query(parquet_path_with_discounts)
    with_combo = run_summary_query(parquet_path_with_discounts, combo_source=[combo_parquet_path])
    assert with_combo["kpis"]["doanhSo"] == without_combo["kpis"]["doanhSo"] == 350000
    assert with_combo["kpis"]["discount"] == without_combo["kpis"]["discount"]
    assert with_combo["kpis"]["voucher"] == without_combo["kpis"]["voucher"]
    assert with_combo["kpis"]["piship"] == without_combo["kpis"]["piship"] == 1620
    assert with_combo["kpis"]["rowCount"] == 3
    assert without_combo["kpis"]["rowCount"] == 2


def test_rows_no_combo_data_leaves_orders_unexploded(parquet_path_with_discounts):
    result = run_rows_query(parquet_path_with_discounts, page_size=10, combo_source=None)
    assert result["total"] == 2
    result_empty_list = run_rows_query(parquet_path_with_discounts, page_size=10, combo_source=[])
    assert result_empty_list["total"] == 2


def test_rows_combo_and_cashflow_together_phi_aff_scales_by_combo_ratio_too(
    parquet_path_with_discounts, combo_parquet_path,
):
    # Phí AFF is prorated by orderPaidRatio first (per the existing Phí AFF
    # feature), then further split by the combo ratio on top of that for an
    # exploded child, per the user's "tất cả các giá trị theo tỉ lệ" rule.
    cashflow_path = _write_cashflow_parquet([{"orderId": "D1", "phiAff": 2000.0}])
    try:
        result = run_rows_query(
            parquet_path_with_discounts, page_size=10,
            combo_source=[combo_parquet_path], cashflow_source=[cashflow_path],
        )
        by_sku = {r["skuVariant"]: r for r in result["rows"]}
        # A100-1's orderPaidRatio is 0.4 (400.000/1.000.000) -> line-level
        # phiAff = 2000*0.4 = 800, then split 0.5/0.5 by combo ratio.
        assert by_sku["X1-1"]["phiAff"] == 400
        assert by_sku["X2-1"]["phiAff"] == 400
        # B200-1: ratio 0.6 -> 2000*0.6 = 1200, no combo match, unscaled.
        assert by_sku["B200-1"]["phiAff"] == 1200
    finally:
        os.remove(cashflow_path)


# ---- Master File cost/category lookup ----
# parquet_path_with_discounts's "D1" order: A100-1 (soLuongThuc=2), B200-1
# (soLuongThuc=1). Master File maps parent SKU "A100" -> Áo/Áo thun/Kho HN/
# giá vốn 10.000, and "B200" -> Quần/Quần short/Kho HCM/giá vốn 20.000.

def _write_master_parquet(rows: list[dict]) -> str:
    return _write_raw_parquet(rows)


@pytest.fixture
def master_parquet_path():
    path = _write_master_parquet([
        {"sku": "A100", "muc": "Áo", "phanLoaiSp": "Áo thun", "phanLoaiKho": "Kho HN", "giaVon": 10000.0},
        {"sku": "B200", "muc": "Quần", "phanLoaiSp": "Quần short", "phanLoaiKho": "Kho HCM", "giaVon": 20000.0},
    ])
    yield path
    os.remove(path)


def test_rows_master_file_lookup_by_parent_sku(parquet_path_with_discounts, master_parquet_path):
    result = run_rows_query(parquet_path_with_discounts, page_size=10, master_source=[master_parquet_path])
    by_sku = {r["skuVariant"]: r for r in result["rows"]}

    assert by_sku["A100-1"]["phanLoaiKho"] == "Kho HN"
    assert by_sku["A100-1"]["phanLoaiMuc"] == "Áo"
    assert by_sku["A100-1"]["phanLoaiSp"] == "Áo thun"
    assert by_sku["A100-1"]["giaVon"] == 2 * 10000  # soLuongThuc x giá vốn

    assert by_sku["B200-1"]["phanLoaiKho"] == "Kho HCM"
    assert by_sku["B200-1"]["giaVon"] == 1 * 20000


def test_rows_combo_children_get_combo_label_but_real_gia_von(
    parquet_path_with_discounts, combo_parquet_path,
):
    # X1/X2 (the exploded children's parent SKUs) have their own Master File
    # entries — Giá vốn must still look them up; the 3 label fields must not.
    master_path = _write_master_parquet([
        {"sku": "X1", "muc": "Áo", "phanLoaiSp": "Áo thun", "phanLoaiKho": "Kho HN", "giaVon": 5000.0},
        {"sku": "X2", "muc": "Quần", "phanLoaiSp": "Quần short", "phanLoaiKho": "Kho HCM", "giaVon": 7000.0},
    ])
    try:
        result = run_rows_query(
            parquet_path_with_discounts, page_size=10,
            combo_source=[combo_parquet_path], master_source=[master_path],
        )
        by_sku = {r["skuVariant"]: r for r in result["rows"]}

        assert by_sku["X1-1"]["phanLoaiKho"] == "combo"
        assert by_sku["X1-1"]["phanLoaiMuc"] == "combo"
        assert by_sku["X1-1"]["phanLoaiSp"] == "combo"
        assert by_sku["X1-1"]["giaVon"] == 2 * 5000  # soLuongThuc (unscaled) x its own giá vốn

        assert by_sku["X2-1"]["phanLoaiKho"] == "combo"
        assert by_sku["X2-1"]["giaVon"] == 2 * 7000
    finally:
        os.remove(master_path)


def test_rows_no_master_match_defaults_to_blank_and_zero(parquet_path_with_discounts, master_parquet_path):
    # "C300-1" (parent "C300") has no Master File entry.
    extra_path = _write_master_parquet([{"sku": "ZZZ", "muc": "X", "phanLoaiSp": "X", "phanLoaiKho": "X", "giaVon": 1.0}])
    try:
        result = run_rows_query(parquet_path_with_discounts, page_size=10, master_source=[extra_path])
        for r in result["rows"]:
            assert r["phanLoaiKho"] == ""
            assert r["phanLoaiMuc"] == ""
            assert r["phanLoaiSp"] == ""
            assert r["giaVon"] == 0
    finally:
        os.remove(extra_path)


def test_rows_no_master_data_uploaded_defaults_same_as_no_match(parquet_path_with_discounts):
    result = run_rows_query(parquet_path_with_discounts, page_size=10, master_source=None)
    for r in result["rows"]:
        assert r["phanLoaiKho"] == ""
        assert r["giaVon"] == 0


def test_rows_duplicate_parent_sku_across_master_rows_does_not_duplicate_orders_rows(
    parquet_path_with_discounts,
):
    # Two Master File rows share parent SKU "A100" (e.g. two colour variants)
    # — ANY_VALUE must pick one consistently, and this must NOT multiply the
    # matching Orders row the way a Combo join intentionally does.
    dup_path = _write_master_parquet([
        {"sku": "A100", "muc": "Áo", "phanLoaiSp": "Áo thun", "phanLoaiKho": "Kho HN", "giaVon": 10000.0},
        {"sku": "A100", "muc": "Áo", "phanLoaiSp": "Áo thun", "phanLoaiKho": "Kho HN", "giaVon": 10000.0},
    ])
    try:
        result = run_rows_query(parquet_path_with_discounts, page_size=10, master_source=[dup_path])
        assert result["total"] == 2  # still exactly 2 Orders rows, not fanned out
    finally:
        os.remove(dup_path)


def test_summary_loi_nhuan_gop_reconciles_with_nmv_and_gia_von(parquet_path_with_discounts, master_parquet_path):
    result = run_summary_query(parquet_path_with_discounts, master_source=[master_parquet_path])
    kpis = result["kpis"]
    assert kpis["giaVon"] == 2 * 10000 + 1 * 20000  # 40000
    assert kpis["loiNhuanGop"] == kpis["nmv"] - kpis["giaVon"]


def test_summary_gia_von_only_counts_gmv_statuses(parquet_path):
    # Same status mix as test_summary_kpis_match_expected: O1 (Hủy sau XK,
    # soLuongThuc 2), O2 (Hủy chưa XK, soLuongThuc 3), O3 (Hoàn hàng,
    # soLuongThuc 0), O4 (Hoàn 1 phần, soLuongThuc 3), O5 (Hoàn thành,
    # soLuongThuc 1), O6 (Đang giao, soLuongThuc 2). Giá vốn must only sum
    # the GMV-status rows (O4, O5, O6), same scope as GMV/discount/voucher —
    # not the cancelled/fully-returned ones.
    master_path = _write_master_parquet([
        {"sku": s, "muc": "X", "phanLoaiSp": "X", "phanLoaiKho": "X", "giaVon": 1000.0}
        for s in ["A100", "B200", "C300", "D400", "E500", "F600"]
    ])
    try:
        result = run_summary_query(parquet_path, master_source=[master_path])
        assert result["kpis"]["giaVon"] == (3 + 1 + 2) * 1000  # 6000, not 11000
    finally:
        os.remove(master_path)


def test_rows_new_filters_narrow_by_master_file_categories_and_sku(
    parquet_path_with_discounts, master_parquet_path,
):
    result_kho = run_rows_query(
        parquet_path_with_discounts, page_size=10, master_source=[master_parquet_path], warehouse_type="Kho HN",
    )
    assert {r["skuVariant"] for r in result_kho["rows"]} == {"A100-1"}

    result_muc = run_rows_query(
        parquet_path_with_discounts, page_size=10, master_source=[master_parquet_path], item_group="Quần",
    )
    assert {r["skuVariant"] for r in result_muc["rows"]} == {"B200-1"}

    result_product_type = run_rows_query(
        parquet_path_with_discounts, page_size=10, master_source=[master_parquet_path], product_type="Áo thun",
    )
    assert {r["skuVariant"] for r in result_product_type["rows"]} == {"A100-1"}

    result_sku = run_rows_query(parquet_path_with_discounts, page_size=10, sku="a100")
    assert {r["skuVariant"] for r in result_sku["rows"]} == {"A100-1"}


# ---- Detail-table "Group theo" sub-tab: run_grouped_rows_query, the
# group_by/group_value drill-down on run_rows_query, and run_export_query.
# Uses parquet_path (O1..O6, categories Áo/Quần, mixed statuses) — doanhSo
# is "originalPrice * quantity" (NOT netted by returns — see
# test_summary_category_filter's comment): O1=200000, O2=150000, O3=80000,
# O4=150000, O5=200000, O6=160000. Áo (O1,O3,O5) sums to 480000, Quần
# (O2,O4,O6) sums to 460000.

def test_grouped_rows_aggregates_match_category_filtered_summary(parquet_path):
    result = run_grouped_rows_query(parquet_path, group_by="category", page_size=10)
    assert result["total"] == 2
    by_group = {r["groupValue"]: r for r in result["rows"]}
    assert set(by_group) == {"Áo", "Quần"}
    assert by_group["Áo"]["rowCount"] == 3
    assert by_group["Quần"]["rowCount"] == 3
    assert by_group["Áo"]["doanhSo"] == run_summary_query(parquet_path, category="Áo")["kpis"]["doanhSo"]
    assert by_group["Quần"]["doanhSo"] == run_summary_query(parquet_path, category="Quần")["kpis"]["doanhSo"]


def test_grouped_rows_sort_by_doanh_so_both_directions(parquet_path):
    desc = run_grouped_rows_query(parquet_path, group_by="category", page_size=10)
    assert [r["groupValue"] for r in desc["rows"]] == ["Áo", "Quần"]  # 480000 > 460000, default desc

    asc = run_grouped_rows_query(parquet_path, group_by="category", sort="doanhSo", sort_dir="asc", page_size=10)
    assert [r["groupValue"] for r in asc["rows"]] == ["Quần", "Áo"]


def test_grouped_rows_pagination(parquet_path):
    page1 = run_grouped_rows_query(parquet_path, group_by="category", page=1, page_size=1)
    page2 = run_grouped_rows_query(parquet_path, group_by="category", page=2, page_size=1)
    assert page1["total"] == page2["total"] == 2
    assert [r["groupValue"] for r in page1["rows"]] == ["Áo"]
    assert [r["groupValue"] for r in page2["rows"]] == ["Quần"]


def test_grouped_rows_respects_existing_filters(parquet_path):
    # Only O5 (Áo) has trangThai exactly "Hoàn thành" (see
    # test_summary_facets_unaffected_by_status_filter).
    result = run_grouped_rows_query(parquet_path, group_by="category", status="Hoàn thành", page_size=10)
    assert result["total"] == 1
    assert result["rows"][0]["groupValue"] == "Áo"
    assert result["rows"][0]["rowCount"] == 1
    assert result["rows"][0]["doanhSo"] == 200000


def test_grouped_rows_group_by_status(parquet_path):
    result = run_grouped_rows_query(parquet_path, group_by="status", page_size=10)
    assert result["total"] == 6
    assert {r["groupValue"] for r in result["rows"]} == {
        "Hủy sau XK", "Hủy chưa XK", "Hoàn hàng", "Hoàn 1 phần", "Hoàn thành", "Đang giao",
    }


def test_grouped_rows_group_by_warehouse_type_from_master_file(parquet_path_with_discounts, master_parquet_path):
    result = run_grouped_rows_query(
        parquet_path_with_discounts, group_by="warehouseType", master_source=[master_parquet_path], page_size=10,
    )
    by_group = {r["groupValue"]: r for r in result["rows"]}
    assert by_group["Kho HN"]["rowCount"] == 1
    assert by_group["Kho HCM"]["rowCount"] == 1


def test_grouped_rows_invalid_group_by_returns_empty(parquet_path):
    result = run_grouped_rows_query(parquet_path, group_by="not_a_real_column", page_size=10)
    assert result == {"rows": [], "total": 0, "page": 1, "pageSize": 10}


def test_rows_drill_down_narrows_to_group(parquet_path):
    result = run_rows_query(parquet_path, page_size=10, path_filters=[("category", "Áo")])
    assert result["total"] == 3
    assert {r["orderId"] for r in result["rows"]} == {"O1", "O3", "O5"}


def test_rows_drill_down_invalid_group_by_ignored(parquet_path):
    result = run_rows_query(parquet_path, page_size=10, path_filters=[("not_a_real_column", "Áo")])
    assert result["total"] == 6  # no filter applied — falls through untouched


def test_rows_drill_down_multi_level_path_filters_all_apply(parquet_path):
    # 2-level path: category="Áo" AND status="Hoàn thành" — only O5 matches
    # both (O1/O3 are Áo but not Hoàn thành).
    result = run_rows_query(
        parquet_path, page_size=10,
        path_filters=[("category", "Áo"), ("status", "Hoàn thành")],
    )
    assert result["total"] == 1
    assert result["rows"][0]["orderId"] == "O5"


def test_grouped_rows_path_filters_narrows_to_nested_ancestor(parquet_path):
    # Nested grouping: level 1 = category, level 2 = status within "Áo".
    # Áo orders: O1 (Hủy sau XK), O3 (Hoàn hàng), O5 (Hoàn thành) — 3 distinct
    # statuses, each its own group of 1.
    result = run_grouped_rows_query(
        parquet_path, group_by="status", page_size=10, path_filters=[("category", "Áo")],
    )
    assert result["total"] == 3
    assert {r["groupValue"] for r in result["rows"]} == {"Hủy sau XK", "Hoàn hàng", "Hoàn thành"}
    assert all(r["rowCount"] == 1 for r in result["rows"])


def test_export_query_ungrouped_returns_every_row_unpaginated(parquet_path):
    rows = run_export_query(parquet_path)
    assert len(rows) == 6
    assert {r["orderId"] for r in rows} == {"O1", "O2", "O3", "O4", "O5", "O6"}


def test_export_query_grouped_returns_every_group_unpaginated(parquet_path):
    rows = run_export_query(parquet_path, group_by="category")
    assert len(rows) == 2
    by_group = {r["groupValue"]: r for r in rows}
    assert by_group["Áo"]["rowCount"] == 3
    assert by_group["Quần"]["rowCount"] == 3


def test_export_query_respects_filters(parquet_path):
    rows = run_export_query(parquet_path, category="Áo")
    assert len(rows) == 3
    assert {r["orderId"] for r in rows} == {"O1", "O3", "O5"}


# ---- Multi-select filters: status/warehouseType/itemGroup/productType each
# now accept a list ("any of these values"), not just one exact match.

def test_rows_status_filter_accepts_list_of_values(parquet_path):
    result = run_rows_query(parquet_path, page_size=10, status=["Hoàn thành", "Đang giao"])
    assert {r["orderId"] for r in result["rows"]} == {"O5", "O6"}


def test_rows_status_filter_still_accepts_single_string(parquet_path):
    result = run_rows_query(parquet_path, page_size=10, status="Hoàn thành")
    assert {r["orderId"] for r in result["rows"]} == {"O5"}


def test_rows_status_filter_empty_list_means_no_filter(parquet_path):
    result = run_rows_query(parquet_path, page_size=10, status=[])
    assert result["total"] == 6


def test_rows_warehouse_type_filter_accepts_list_of_values(parquet_path_with_discounts, master_parquet_path):
    result = run_rows_query(
        parquet_path_with_discounts, page_size=10, master_source=[master_parquet_path],
        warehouse_type=["Kho HN", "Kho HCM"],
    )
    assert {r["skuVariant"] for r in result["rows"]} == {"A100-1", "B200-1"}

    result_one = run_rows_query(
        parquet_path_with_discounts, page_size=10, master_source=[master_parquet_path],
        warehouse_type=["Kho HN"],
    )
    assert {r["skuVariant"] for r in result_one["rows"]} == {"A100-1"}


def test_summary_status_filter_accepts_list_of_values(parquet_path):
    result = run_summary_query(parquet_path, status=["Hoàn thành", "Đang giao"])
    assert result["kpis"]["rowCount"] == 2


# ---- Per-row GMV/Doanh thu thuần/NMV/Lợi nhuận gộp columns (Detail-table
# "Cột hiển thị") — must reconcile with the KPI cards when summed, since
# they use the exact same GMV-status scoping as run_summary_query's totals.

def test_rows_gmv_column_excludes_non_gmv_statuses(parquet_path):
    rows = run_export_query(parquet_path)
    by_order = {r["orderId"]: r for r in rows}
    assert by_order["O1"]["gmv"] == 0  # Hủy sau XK
    assert by_order["O2"]["gmv"] == 0  # Hủy chưa XK
    assert by_order["O5"]["gmv"] == 200000  # Hoàn thành: originalPrice(200000) x soLuongThuc(1)


def test_rows_gmv_nmv_loi_nhuan_gop_columns_reconcile_with_summary(
    parquet_path_with_discounts, master_parquet_path,
):
    summary = run_summary_query(parquet_path_with_discounts, master_source=[master_parquet_path])
    rows = run_export_query(parquet_path_with_discounts, master_source=[master_parquet_path])
    assert sum(r["gmv"] for r in rows) == summary["kpis"]["gmv"]
    assert sum(r["doanhThuThuan"] for r in rows) == summary["kpis"]["doanhThuThuan"]
    assert sum(r["nmv"] for r in rows) == summary["kpis"]["nmv"]
    assert sum(r["loiNhuanGop"] for r in rows) == summary["kpis"]["loiNhuanGop"]


def test_grouped_rows_gmv_nmv_loi_nhuan_gop_reconcile_with_summary(
    parquet_path_with_discounts, master_parquet_path,
):
    result = run_grouped_rows_query(
        parquet_path_with_discounts, group_by="sku", master_source=[master_parquet_path], page_size=10,
    )
    summary = run_summary_query(parquet_path_with_discounts, master_source=[master_parquet_path])
    assert sum(r["gmv"] for r in result["rows"]) == summary["kpis"]["gmv"]
    assert sum(r["nmv"] for r in result["rows"]) == summary["kpis"]["nmv"]
    assert sum(r["loiNhuanGop"] for r in result["rows"]) == summary["kpis"]["loiNhuanGop"]


# ---- Sales Channel ("Kênh bán hàng") join — each Orders Report's parquet
# path is grouped by channel name (channel_source: {name: [paths]}) since
# there's no per-row marker for which Report a row came from otherwise.
# Written to CONTROLLED filenames (not tempfile.mkstemp's random name) so
# channel_source's path lists can reference them directly.

def _write_orders_parquet_at(path: str, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    parquet_bytes, row_count, _ = excel_to_parquet(buf)
    assert row_count == len(rows)
    with open(path, "wb") as f:
        f.write(parquet_bytes)


@pytest.fixture
def channel_tagged_reports():
    tmpdir = tempfile.mkdtemp()
    shopee_path = os.path.join(tmpdir, "R1.parquet")
    lazada_path = os.path.join(tmpdir, "R2.parquet")
    untagged_path = os.path.join(tmpdir, "R3.parquet")
    _write_orders_parquet_at(shopee_path, [
        ["S1", "2026-03-01 00:00", "Hoàn thành", "", "A100-1", "SP A", "Áo", 100000, 1, 0],
    ])
    _write_orders_parquet_at(lazada_path, [
        ["L1", "2026-03-02 00:00", "Hoàn thành", "", "B200-1", "SP B", "Quần", 50000, 1, 0],
    ])
    _write_orders_parquet_at(untagged_path, [
        ["U1", "2026-03-03 00:00", "Hoàn thành", "", "C300-1", "SP C", "Áo", 20000, 1, 0],
    ])
    yield {"shopee": shopee_path, "lazada": lazada_path, "untagged": untagged_path}
    for p in (shopee_path, lazada_path, untagged_path):
        os.remove(p)
    os.rmdir(tmpdir)


def test_rows_sales_channel_tags_rows_by_report(channel_tagged_reports):
    paths = channel_tagged_reports
    all_paths = [paths["shopee"], paths["lazada"], paths["untagged"]]
    channel_source = {"Shopee": [paths["shopee"]], "Lazada": [paths["lazada"]]}
    result = run_rows_query(all_paths, page_size=10, channel_source=channel_source)
    by_order = {r["orderId"]: r for r in result["rows"]}
    assert by_order["S1"]["salesChannel"] == "Shopee"
    assert by_order["L1"]["salesChannel"] == "Lazada"
    assert by_order["U1"]["salesChannel"] == ""  # not in any channel group -> unassigned


def test_rows_no_channel_source_defaults_all_rows_unassigned(channel_tagged_reports):
    paths = channel_tagged_reports
    all_paths = [paths["shopee"], paths["lazada"], paths["untagged"]]
    result = run_rows_query(all_paths, page_size=10, channel_source=None)
    assert all(r["salesChannel"] == "" for r in result["rows"])


def test_summary_sales_channel_facet_lists_every_channel(channel_tagged_reports):
    paths = channel_tagged_reports
    all_paths = [paths["shopee"], paths["lazada"], paths["untagged"]]
    channel_source = {"Shopee": [paths["shopee"]], "Lazada": [paths["lazada"]]}
    result = run_summary_query(all_paths, channel_source=channel_source)
    assert set(result["facets"]["salesChannels"]) == {"Shopee", "Lazada"}  # "" filtered out, like other facets


def test_rows_sales_channel_filter_narrows_to_matching_reports(channel_tagged_reports):
    paths = channel_tagged_reports
    all_paths = [paths["shopee"], paths["lazada"], paths["untagged"]]
    channel_source = {"Shopee": [paths["shopee"]], "Lazada": [paths["lazada"]]}
    result = run_rows_query(all_paths, page_size=10, channel_source=channel_source, sales_channel=["Shopee"])
    assert {r["orderId"] for r in result["rows"]} == {"S1"}


def test_grouped_rows_by_sales_channel(channel_tagged_reports):
    paths = channel_tagged_reports
    all_paths = [paths["shopee"], paths["lazada"], paths["untagged"]]
    channel_source = {"Shopee": [paths["shopee"]], "Lazada": [paths["lazada"]]}
    result = run_grouped_rows_query(
        all_paths, group_by="salesChannel", channel_source=channel_source, page_size=10,
    )
    by_group = {r["groupValue"]: r for r in result["rows"]}
    assert set(by_group) == {"Shopee", "Lazada", ""}
    assert by_group["Shopee"]["rowCount"] == 1
    assert by_group[""]["rowCount"] == 1  # the untagged Report


def test_channel_override_wins_over_report_level_channel_tag():
    # The combined 31 LVS/HARA/WEBSITE/ZALO file (see
    # excel_to_parquet.build_dashboard_rows) stores each row's real Kênh
    # bán hàng in "channelOverride" — this must win over whatever channel
    # (or lack of one) the Report itself was tagged with at upload time,
    # since one Report can genuinely mix several of these 4 channels.
    rows = [
        {"date": datetime(2026, 2, 1), "orderId": "C1", "sku": "X1",
         "skuVariant": "X1-1", "product": "SP X", "category": "Áo", "customer": "(Không rõ)",
         "quantity": 1.0, "returnedQty": 0.0, "soLuongThuc": 1.0, "price": 0.0, "originalPrice": 10000.0,
         "revenue": 0.0, "doanhSo": 10000.0, "status": "Hoàn thành", "trangThai": "Hoàn thành",
         "channelOverride": "HARA"},
        {"date": datetime(2026, 2, 1), "orderId": "C2", "sku": "X1",
         "skuVariant": "X1-1", "product": "SP X", "category": "Áo", "customer": "(Không rõ)",
         "quantity": 1.0, "returnedQty": 0.0, "soLuongThuc": 1.0, "price": 0.0, "originalPrice": 10000.0,
         "revenue": 0.0, "doanhSo": 10000.0, "status": "Hoàn thành", "trangThai": "Hoàn thành",
         "channelOverride": ""},  # blank -> falls back to the Report-level tag
    ]
    path = _write_raw_parquet(rows)
    try:
        # Report-level tag says "Shopee" — only C2 (blank override) should
        # inherit it; C1's "HARA" override must win regardless.
        result = run_rows_query([path], page_size=10, channel_source={"Shopee": [path]})
        by_order = {r["orderId"]: r["salesChannel"] for r in result["rows"]}
        assert by_order == {"C1": "HARA", "C2": "Shopee"}
    finally:
        os.remove(path)


# ---- "Kênh nhỏ" (LIVE/VIDEO/PSA/AFF) classification ----
# Confirmed with the user 2026-08-27: TikTok-only; a Kênh AFF file match
# always wins (AFF); a blank/"0" Creator Handle is the main channel (PSA);
# an admin-managed "ID Inhouse" handle maps by Order Channel; any other
# handle is an outside creator (AFF); every other sales channel stays NULL.

def _kenh_nho_row(order_id: str, sku_id: str, creator_handle, content_channel) -> dict:
    return {
        "date": datetime(2026, 2, 1), "orderId": order_id, "sku": "X1",
        "skuVariant": "X1-1", "product": "SP X", "category": "Áo", "customer": "(Không rõ)",
        "quantity": 1.0, "returnedQty": 0.0, "soLuongThuc": 1.0, "price": 0.0, "originalPrice": 10000.0,
        "revenue": 0.0, "doanhSo": 10000.0, "status": "Hoàn thành", "trangThai": "Hoàn thành",
        "discount": 0.0, "voucher": 0.0, "skuId": sku_id,
        "creatorHandle": creator_handle, "contentChannel": content_channel,
    }


@pytest.fixture
def tiktok_kenh_nho_path():
    rows = [
        _kenh_nho_row("O1", "S1", "", ""),  # blank handle -> PSA
        _kenh_nho_row("O2", "S2", "0", "Videos"),  # "0" handle -> PSA (Order Channel ignored)
        _kenh_nho_row("O3", "S3", "bbstores.vn", "Videos"),  # inhouse -> VIDEO
        _kenh_nho_row("O4", "S4", "bbstores.vn", "LIVE"),  # inhouse -> LIVE
        _kenh_nho_row("O5", "S5", "BBCongSo", "Product cards"),  # inhouse, case-insensitive -> PSA
        _kenh_nho_row("O6", "S6", "randomcreator123", "Videos"),  # outside creator -> AFF
        _kenh_nho_row("O7", "S7", "bbstores.vn", "Videos"),  # matched in Kênh AFF file -> AFF wins over inhouse
    ]
    path = _write_raw_parquet(rows)
    yield path
    os.remove(path)


@pytest.fixture
def aff_channel_path():
    path = _write_raw_parquet([{"orderId": "O7", "skuId": "S7"}])
    yield path
    os.remove(path)


INHOUSE_HANDLES = ["bbstores.vn", "bbcongso", "bbstores_forlady"]


def test_kenh_nho_classification_all_branches(tiktok_kenh_nho_path, aff_channel_path):
    channel_source = {"TikTok": [tiktok_kenh_nho_path]}
    result = run_rows_query(
        [tiktok_kenh_nho_path], page_size=20, channel_source=channel_source,
        aff_source=[aff_channel_path], inhouse_handles=INHOUSE_HANDLES,
    )
    by_order = {r["orderId"]: r["kenhNho"] for r in result["rows"]}
    assert by_order == {
        "O1": "PSA", "O2": "PSA", "O3": "VIDEO", "O4": "LIVE",
        "O5": "PSA", "O6": "AFF", "O7": "AFF",
    }


def test_kenh_nho_stays_null_for_non_tiktok_channel(tiktok_kenh_nho_path, aff_channel_path):
    channel_source = {"Shopee": [tiktok_kenh_nho_path]}
    result = run_rows_query(
        [tiktok_kenh_nho_path], page_size=20, channel_source=channel_source,
        aff_source=[aff_channel_path], inhouse_handles=INHOUSE_HANDLES,
    )
    assert all(r["kenhNho"] is None for r in result["rows"])


def test_kenh_nho_null_when_untagged_and_no_inhouse_or_aff_data(tiktok_kenh_nho_path):
    # No channel_source at all -> salesChannel defaults to "" (not "tiktok"),
    # so kenhNho stays NULL even though creatorHandle/contentChannel exist.
    result = run_rows_query([tiktok_kenh_nho_path], page_size=20)
    assert all(r["kenhNho"] is None for r in result["rows"])


def test_kenh_nho_facet_lists_only_non_null_values(tiktok_kenh_nho_path, aff_channel_path):
    channel_source = {"TikTok": [tiktok_kenh_nho_path]}
    result = run_summary_query(
        [tiktok_kenh_nho_path], channel_source=channel_source,
        aff_source=[aff_channel_path], inhouse_handles=INHOUSE_HANDLES,
    )
    assert set(result["facets"]["kenhNho"]) == {"PSA", "VIDEO", "LIVE", "AFF"}


def test_kenh_nho_filter_narrows_rows(tiktok_kenh_nho_path, aff_channel_path):
    channel_source = {"TikTok": [tiktok_kenh_nho_path]}
    result = run_rows_query(
        [tiktok_kenh_nho_path], page_size=20, channel_source=channel_source,
        aff_source=[aff_channel_path], inhouse_handles=INHOUSE_HANDLES, kenh_nho=["AFF"],
    )
    assert {r["orderId"] for r in result["rows"]} == {"O6", "O7"}


def test_kenh_nho_old_schema_report_without_new_columns_defaults_to_psa_when_tiktok():
    # A TikTok Report converted before skuId/creatorHandle/contentChannel
    # existed: the columns are globally absent (not just NULL per-row) —
    # must not error, and best-effort classifies as PSA (same as a real
    # blank Creator Handle) rather than crashing the whole Dashboard.
    rows = [{
        "date": datetime(2026, 2, 1), "orderId": "OLD1", "sku": "X1",
        "skuVariant": "X1-1", "product": "SP X", "category": "Áo", "customer": "(Không rõ)",
        "quantity": 1.0, "returnedQty": 0.0, "soLuongThuc": 1.0, "price": 0.0, "originalPrice": 10000.0,
        "revenue": 0.0, "doanhSo": 10000.0, "status": "Hoàn thành", "trangThai": "Hoàn thành",
    }]
    path = _write_raw_parquet(rows)
    try:
        result = run_rows_query([path], page_size=10, channel_source={"TikTok": [path]})
        assert result["rows"][0]["kenhNho"] == "PSA"
    finally:
        os.remove(path)
